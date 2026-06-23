import io
import logging
from datetime import date, datetime
from decimal import Decimal
from typing import Optional

from fastapi import APIRouter, UploadFile, File, HTTPException
from sqlalchemy import text
from sqlalchemy.ext.asyncio import AsyncSession

from app.database import get_db
from fastapi import Depends

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/truestat", tags=["upload"])


# ======================================================================
# Статистика продаж → product_daily_stats
# ======================================================================

PDS_COLUMN_MAP = {
    "nm": "nm_id", "нм": "nm_id", "nm_id": "nm_id", "nm id": "nm_id",
    "дата": "date", "date": "date", "дата продажи": "date",
    "заказы": "order_count", "заказы всего": "order_count", "заказы, шт": "order_count",
    "order_count": "order_count",
    "сумма заказов": "order_sum", "заказы на сумму": "order_sum",
    "order_sum": "order_sum",
    "выкупы": "buyout_count", "выкупили": "buyout_count", "выкупили, шт": "buyout_count",
    "buyout_count": "buyout_count",
    "сумма выкупов": "buyout_sum", "выкупили на сумму": "buyout_sum",
    "buyout_sum": "buyout_sum",
    "% выкупа": "buyout_percent", "процент выкупа": "buyout_percent",
    "buyout_percent": "buyout_percent",
    "возвраты": "cancel_count", "отмены": "cancel_count", "отменили": "cancel_count",
    "cancel_count": "cancel_count",
    "сумма возвратов": "cancel_sum", "отмены на сумму": "cancel_sum",
    "cancel_sum": "cancel_sum",
    "реклама": "ad_spend", "расходы на рекламу": "ad_spend",
    "ad_spend": "ad_spend",
    "цена до скидки": "avg_price_before_spp",
    "avg_price_before_spp": "avg_price_before_spp",
    "цена после скидки": "avg_price_after_spp",
    "avg_price_after_spp": "avg_price_after_spp",
    "скидка %": "avg_spp_pct", "spp%": "avg_spp_pct", "avg_spp_pct": "avg_spp_pct",
    "просмотры": "open_count", "open_count": "open_count",
    "в корзину": "cart_count", "cart_count": "cart_count",
}

# ======================================================================
# Отчёт о реализации (фин отчёт) → realization_daily_stats
# ======================================================================

REAL_COLUMN_MAP = {
    "nm": "nm_id", "нм": "nm_id", "nm_id": "nm_id",
    "дата документа": "rr_dt", "дата операции": "rr_dt",
    "rr_dt": "rr_dt",
    "операция": "supplier_oper_name", "тип операции": "supplier_oper_name",
    "supplier_oper_name": "supplier_oper_name",
    "количество": "quantity", "quantity": "quantity", "кол-во": "quantity",
    "цена розничная": "retail_price_withdisc_rub",
    "retail_price_withdisc_rub": "retail_price_withdisc_rub",
    "доставка": "delivery_rub", "delivery_rub": "delivery_rub",
    "стоимость хранения": "storage_fee", "storage_fee": "storage_fee",
    "приёмка": "acceptance", "acceptance": "acceptance",
    "удержания": "deduction", "deduction": "deduction",
    "штрафы": "penalty", "penalty": "penalty",
    "доплаты": "additional_payment", "additional_payment": "additional_payment",
    "эквайринг": "acquiring_fee", "acquiring_fee": "acquiring_fee",
    "комиссия": "ppvz_sales_commission",
    "ppvz_sales_commission": "ppvz_sales_commission",
    "к выплате": "ppvz_for_pay", "ppvz_for_pay": "ppvz_for_pay",
}


# ======================================================================
# Helpers
# ======================================================================

def _to_float(v, default=0.0) -> float:
    if v is None:
        return default
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, Decimal):
        return float(v)
    s = str(v).strip().replace(" ", "").replace(",", ".").replace("₽", "").replace("%", "")
    if not s:
        return default
    try:
        return float(s)
    except ValueError:
        return default


def _to_int(v) -> int:
    return int(_to_float(v))


def _parse_date(v):
    if isinstance(v, date):
        return v
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, str):
        try:
            return date.fromisoformat(v[:10])
        except ValueError:
            pass
        fmts = ["%d.%m.%Y", "%Y-%m-%d", "%d/%m/%Y"]
        for f in fmts:
            try:
                return datetime.strptime(v[:10].strip(), f).date()
            except ValueError:
                continue
    return None


def _normalize_headers(headers, col_map):
    mapping = {}
    for i, h in enumerate(headers):
        key = str(h).strip().lower()
        for repl in ["\n", "\r", "  "]:
            while repl in key:
                key = key.replace(repl, " ")
        key = key.strip()
        mapped = col_map.get(key)
        if mapped:
            mapping[i] = mapped
    return mapping


# ======================================================================
# Import: product_daily_stats
# ======================================================================

PDS_UPSERT = text("""
    INSERT INTO product_daily_stats (
        nm_id, date, open_count, cart_count,
        order_count, order_sum, buyout_count, buyout_sum, buyout_percent,
        cancel_count, cancel_sum, ad_spend,
        avg_price_before_spp, avg_price_after_spp, avg_spp_pct
    ) VALUES (
        :nm_id, :date, :open_count, :cart_count,
        :order_count, :order_sum, :buyout_count, :buyout_sum, :buyout_percent,
        :cancel_count, :cancel_sum, :ad_spend,
        :avg_price_before_spp, :avg_price_after_spp, :avg_spp_pct
    )
    ON CONFLICT (nm_id, date) DO UPDATE SET
        open_count = CASE WHEN EXCLUDED.open_count > 0 THEN EXCLUDED.open_count ELSE product_daily_stats.open_count END,
        cart_count = CASE WHEN EXCLUDED.cart_count > 0 THEN EXCLUDED.cart_count ELSE product_daily_stats.cart_count END,
        order_count = CASE WHEN EXCLUDED.order_count > 0 THEN EXCLUDED.order_count ELSE product_daily_stats.order_count END,
        order_sum = CASE WHEN EXCLUDED.order_sum > 0 THEN EXCLUDED.order_sum ELSE product_daily_stats.order_sum END,
        buyout_count = CASE WHEN EXCLUDED.buyout_count > 0 THEN EXCLUDED.buyout_count ELSE product_daily_stats.buyout_count END,
        buyout_sum = CASE WHEN EXCLUDED.buyout_sum > 0 THEN EXCLUDED.buyout_sum ELSE product_daily_stats.buyout_sum END,
        buyout_percent = CASE WHEN EXCLUDED.buyout_percent > 0 THEN EXCLUDED.buyout_percent ELSE product_daily_stats.buyout_percent END,
        cancel_count = CASE WHEN EXCLUDED.cancel_count > 0 THEN EXCLUDED.cancel_count ELSE product_daily_stats.cancel_count END,
        cancel_sum = CASE WHEN EXCLUDED.cancel_sum > 0 THEN EXCLUDED.cancel_sum ELSE product_daily_stats.cancel_sum END,
        ad_spend = CASE WHEN EXCLUDED.ad_spend > 0 THEN EXCLUDED.ad_spend ELSE product_daily_stats.ad_spend END,
        avg_price_before_spp = CASE WHEN EXCLUDED.avg_price_before_spp > 0 THEN EXCLUDED.avg_price_before_spp ELSE product_daily_stats.avg_price_before_spp END,
        avg_price_after_spp = CASE WHEN EXCLUDED.avg_price_after_spp > 0 THEN EXCLUDED.avg_price_after_spp ELSE product_daily_stats.avg_price_after_spp END,
        avg_spp_pct = CASE WHEN EXCLUDED.avg_spp_pct > 0 THEN EXCLUDED.avg_spp_pct ELSE product_daily_stats.avg_spp_pct END
""")


async def _import_pds(db: AsyncSession, rows: list[dict]) -> dict:
    imported = 0
    for row in rows:
        nm_id = _to_int(row.get("nm_id", 0))
        if not nm_id:
            continue
        dt = _parse_date(row.get("date"))
        if dt is None:
            continue

        params = {
            "nm_id": nm_id,
            "date": dt,
            "open_count": _to_int(row.get("open_count", 0)),
            "cart_count": _to_int(row.get("cart_count", 0)),
            "order_count": _to_int(row.get("order_count", 0)),
            "order_sum": _to_float(row.get("order_sum", 0)),
            "buyout_count": _to_int(row.get("buyout_count", 0)),
            "buyout_sum": _to_float(row.get("buyout_sum", 0)),
            "buyout_percent": _to_float(row.get("buyout_percent", 0)),
            "cancel_count": _to_int(row.get("cancel_count", 0)),
            "cancel_sum": _to_float(row.get("cancel_sum", 0)),
            "ad_spend": _to_float(row.get("ad_spend", 0)),
            "avg_price_before_spp": _to_float(row.get("avg_price_before_spp", 0)),
            "avg_price_after_spp": _to_float(row.get("avg_price_after_spp", 0)),
            "avg_spp_pct": _to_float(row.get("avg_spp_pct", 0)),
        }
        if any(v > 0 for v in [params["order_count"], params["order_sum"],
                               params["ad_spend"], params["avg_price_before_spp"]]):
            await db.execute(PDS_UPSERT, params)
            imported += 1

    await db.commit()
    recalc = await _recalc_roi(db)
    return {"imported_pds": imported, "recalc_roi": recalc}


# ======================================================================
# Import: realization_daily_stats (фин отчёт)
# ======================================================================

REAL_DELETE_SQL = text("DELETE FROM realization_daily_stats WHERE stat_date BETWEEN :start AND :end")

REAL_UPSERT = text("""
    INSERT INTO realization_daily_stats (
        stat_date, nm_id, sales_count, sales_revenue, returns_count, returns_revenue,
        logistics_cost, rebill_logistics_cost, storage_cost, acceptance_cost,
        deduction_cost, penalty_cost, additional_payment, acquiring_sales,
        acquiring_returns, commission_sales, commission_returns, payout_sales,
        payout_returns, net_payout, net_qty, cost_price_estimate, profit_estimate, updated_at
    ) VALUES (
        :stat_date, :nm_id, :sales_count, :sales_revenue, :returns_count, :returns_revenue,
        :logistics_cost, :rebill_logistics_cost, :storage_cost, :acceptance_cost,
        :deduction_cost, :penalty_cost, :additional_payment, :acquiring_sales,
        :acquiring_returns, :commission_sales, :commission_returns, :payout_sales,
        :payout_returns, :net_payout, :net_qty, :cost_price_estimate, :profit_estimate, now()
    )
    ON CONFLICT (stat_date, nm_id) DO UPDATE SET
        sales_count = EXCLUDED.sales_count,
        sales_revenue = EXCLUDED.sales_revenue,
        returns_count = EXCLUDED.returns_count,
        returns_revenue = EXCLUDED.returns_revenue,
        logistics_cost = EXCLUDED.logistics_cost,
        storage_cost = EXCLUDED.storage_cost,
        net_payout = EXCLUDED.net_payout,
        cost_price_estimate = EXCLUDED.cost_price_estimate,
        profit_estimate = EXCLUDED.profit_estimate,
        updated_at = now()
""")


async def _import_realization(db: AsyncSession, rows: list[dict]) -> dict:
    agg = {}
    date_min = None
    date_max = None

    cost_sql = await db.execute(text("SELECT nm_id, cost_price FROM products"))
    cost_map = {r[0]: float(r[1] or 0) for r in cost_sql.fetchall()}

    for row in rows:
        rr_dt = _parse_date(row.get("rr_dt"))
        if rr_dt is None:
            continue
        nm_id = _to_int(row.get("nm_id", 0))
        if not nm_id:
            continue

        if date_min is None or rr_dt < date_min:
            date_min = rr_dt
        if date_max is None or rr_dt > date_max:
            date_max = rr_dt

        key = (rr_dt, nm_id)
        item = agg.setdefault(key, {
            "sales_count": 0, "sales_revenue": 0.0,
            "returns_count": 0, "returns_revenue": 0.0,
            "logistics_cost": 0.0, "rebill_logistics_cost": 0.0,
            "storage_cost": 0.0, "acceptance_cost": 0.0,
            "deduction_cost": 0.0, "penalty_cost": 0.0,
            "additional_payment": 0.0,
            "acquiring_sales": 0.0, "acquiring_returns": 0.0,
            "commission_sales": 0.0, "commission_returns": 0.0,
            "payout_sales": 0.0, "payout_returns": 0.0,
        })

        op = str(row.get("supplier_oper_name", "") or "")
        qty = _to_int(row.get("quantity", 0)) or 1
        retail = _to_float(row.get("retail_price_withdisc_rub", 0))
        delivery = _to_float(row.get("delivery_rub", 0))
        rebill = _to_float(row.get("rebill_logistic_cost", 0))
        storage = _to_float(row.get("storage_fee", 0))
        acceptance = _to_float(row.get("acceptance", 0))
        deduction = _to_float(row.get("deduction", 0))
        penalty = _to_float(row.get("penalty", 0))
        add_pay = _to_float(row.get("additional_payment", 0))
        acquiring = _to_float(row.get("acquiring_fee", 0))
        commission = abs(_to_float(row.get("ppvz_sales_commission", 0)))
        payout = _to_float(row.get("ppvz_for_pay", 0))

        if "продаж" in op.lower():
            item["sales_count"] += qty
            item["sales_revenue"] += retail
            item["acquiring_sales"] += acquiring
            item["commission_sales"] += commission
            item["payout_sales"] += payout
        elif "возврат" in op.lower():
            item["returns_count"] += qty
            item["returns_revenue"] += retail
            item["acquiring_returns"] += acquiring
            item["commission_returns"] += commission
            item["payout_returns"] += payout
        elif "логистик" in op.lower():
            item["logistics_cost"] += delivery
        elif "хранен" in op.lower():
            item["storage_cost"] += storage

        item["rebill_logistics_cost"] += rebill
        item["acceptance_cost"] += acceptance
        item["deduction_cost"] += deduction
        item["penalty_cost"] += penalty
        item["additional_payment"] += add_pay

    if date_min is not None and date_max is not None:
        await db.execute(REAL_DELETE_SQL, {"start": date_min, "end": date_max})

    count = 0
    for (stat_date, nm_id), item in agg.items():
        net_qty = item["sales_count"] - item["returns_count"]
        cost_price_estimate = cost_map.get(nm_id, 0.0) * net_qty
        net_payout = (
            item["payout_sales"] - item["payout_returns"]
            - item["logistics_cost"] - item["rebill_logistics_cost"]
            - item["storage_cost"] - item["acceptance_cost"]
            - item["deduction_cost"] - item["penalty_cost"]
            + item["additional_payment"]
        )
        profit_estimate = net_payout - cost_price_estimate

        params = {
            "stat_date": stat_date,
            "nm_id": nm_id,
            **item,
            "net_payout": net_payout,
            "net_qty": net_qty,
            "cost_price_estimate": cost_price_estimate,
            "profit_estimate": profit_estimate,
        }
        await db.execute(REAL_UPSERT, params)
        count += 1

    await db.commit()
    return {"imported_real": count, "date_from": str(date_min), "date_to": str(date_max)}


# ======================================================================
# Custom format: two-sheet file (vendor code based)
# Sheet 1: sales summary by vendor code + date
# Sheet 2: ad spend by vendor code + date
# ======================================================================

REPORT_HEADERS_SHEET1 = {"артикул продавца", "выкупили, шт.", "заказано, шт."}
REPORT_HEADERS_SHEET2 = {"артикул", "расход, \u20bd", "день"}


async def _detect_and_import_report(db: AsyncSession, data: bytes) -> dict:
    try:
        import openpyxl
    except ImportError:
        raise HTTPException(500, "openpyxl not installed")

    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    if len(wb.sheetnames) < 1:
        raise HTTPException(400, "Empty workbook")

    ws1 = wb[wb.sheetnames[0]]
    h1 = [str(h).strip().lower() if h else "" for h in next(ws1.iter_rows(values_only=True))]
    h1_set = set(h1)

    if not (REPORT_HEADERS_SHEET1 & h1_set):
        return None  # not our format

    # Map sheet 1 columns
    col_idx = {v: i for i, v in enumerate(h1)}
    date_col = col_idx.get("", 2)  # date is in unnamed column
    vc_col = col_idx.get("артикул продавца", 3)
    buyout_col = col_idx.get("выкупили, шт.", 4)
    payout_col = col_idx.get("к перечислению за товар, руб.", 5)
    order_col = col_idx.get("заказано, шт.", 6)
    order_sum_col = col_idx.get("сумма заказов минус комиссия wb, руб.", 7)

    # Load vendor_code → nm_id mapping
    vc_rows = await db.execute(text("SELECT vendor_code, nm_id FROM products"))
    vc_map = {str(r[0]).strip().lower(): r[1] for r in vc_rows.fetchall() if r[0]}

    # Parse sheet 1
    sheet1_data = {}
    for row in ws1.iter_rows(values_only=True):
        vals = list(row)
        if not vals or not vals[vc_col]:
            continue
        vc = str(vals[vc_col]).strip()
        vc_lower = vc.lower()
        nm_id = vc_map.get(vc_lower)
        if not nm_id:
            continue

        raw_date = vals[date_col]
        dt = None
        if isinstance(raw_date, datetime):
            dt = raw_date.date()
        elif isinstance(raw_date, date):
            dt = raw_date
        elif isinstance(raw_date, str):
            try:
                dt = datetime.strptime(raw_date[:10], "%Y-%m-%d").date()
            except ValueError:
                continue
        if dt is None:
            continue

        key = (nm_id, dt)
        item = sheet1_data.setdefault(key, {
            "order_count": 0, "order_sum": 0.0,
            "buyout_count": 0, "buyout_sum": 0.0,
        })
        item["order_count"] += _to_int(vals[order_col]) if vals[order_col] else 0
        item["order_sum"] += _to_float(vals[order_sum_col]) if order_sum_col < len(vals) and vals[order_sum_col] else 0
        item["buyout_count"] += _to_int(vals[buyout_col]) if vals[buyout_col] else 0
        item["buyout_sum"] += _to_float(vals[payout_col]) if vals[payout_col] else 0

    # Try sheet 2 (ad spend)
    ws2 = wb[wb.sheetnames[1]] if len(wb.sheetnames) > 1 else None
    ad_data = {}
    if ws2:
        h2_raw = next(ws2.iter_rows(values_only=True))
        h2 = [str(h).strip().lower() if h else "" for h in h2_raw]
        has_ad = "расход" in str(h2) or "расход без бонусов" in str(h2)
        if has_ad:
            vc2_col = next((i for i, h in enumerate(h2) if h == "артикул"), 0)
            date2_col = next((i for i, h in enumerate(h2) if "день" in h or "дата" in h), 4)
            spend_col = next((i for i, h in enumerate(h2) if "расход" in h and "руб" in h), 12)
            if spend_col >= len(h2):
                spend_col = next((i for i, h in enumerate(h2) if "расход" in h), 5)

            for row in ws2.iter_rows(values_only=True):
                vals = list(row)
                if not vals or not vals[vc2_col]:
                    continue
                vc = str(vals[vc2_col]).strip().lower()
                nm_id = vc_map.get(vc)
                if not nm_id:
                    continue

                raw_date = vals[date2_col]
                dt = None
                if isinstance(raw_date, datetime):
                    dt = raw_date.date()
                elif isinstance(raw_date, date):
                    dt = raw_date
                elif isinstance(raw_date, str):
                    try:
                        dt = datetime.strptime(raw_date[:10], "%Y-%m-%d").date()
                    except ValueError:
                        continue
                if dt is None:
                    continue

                key = (nm_id, dt)
                spend = _to_float(vals[spend_col]) if spend_col < len(vals) and vals[spend_col] else 0
                ad_data[key] = ad_data.get(key, 0) + spend

    # Merge ad data
    for key, spend in ad_data.items():
        if key in sheet1_data:
            sheet1_data[key]["ad_spend"] = sheet1_data[key].get("ad_spend", 0) + spend
        else:
            sheet1_data[key] = {
                "order_count": 0, "order_sum": 0.0,
                "buyout_count": 0, "buyout_sum": 0.0,
                "ad_spend": spend,
            }

    # Import
    imported = 0
    upsert = PDS_UPSERT  # same upsert SQL
    for (nm_id, dt), item in sheet1_data.items():
        params = {
            "nm_id": nm_id, "date": dt,
            "open_count": 0, "cart_count": 0,
            "order_count": item["order_count"],
            "order_sum": item["order_sum"],
            "buyout_count": item["buyout_count"],
            "buyout_sum": item["buyout_sum"],
            "buyout_percent": 0,
            "cancel_count": 0, "cancel_sum": 0,
            "ad_spend": item.get("ad_spend", 0),
            "avg_price_before_spp": 0,
            "avg_price_after_spp": 0,
            "avg_spp_pct": 0,
        }
        if params["order_count"] > 0 or params["ad_spend"] > 0:
            await db.execute(upsert, params)
            imported += 1

    await db.commit()

    # Auto-recalculate ROI after import
    recalc_rows = await _recalc_roi(db)
    return {"imported": imported, "ad_rows": len(ad_data), "recalc_roi": recalc_rows}


# ======================================================================
# ROI recalculation after import
# ======================================================================

async def _recalc_roi(db: AsyncSession, since_date: date = date(2026, 1, 1)) -> int:
    from app.services.logistics_calc import calc_metrics

    def _f(v):
        if isinstance(v, Decimal): return float(v)
        return float(v or 0)

    r = await db.execute(text("""
        SELECT pds.id, pds.order_count, pds.order_sum,
               pds.buyout_count, pds.buyout_sum, pds.buyout_percent,
               pds.ad_spend, p.cost_price, p.volume_liters,
               p.warehouse_coef, p.buyout_percent as product_bp
        FROM product_daily_stats pds
        JOIN products p ON p.nm_id = p.nm_id
        WHERE pds.date >= :since AND pds.order_count > 0
    """), {"since": since_date})

    updated = 0
    for row in r.fetchall():
        bp = _f(row.product_bp) or _f(row.buyout_percent) or 20
        if bp < 5: bp = 20
        metrics = calc_metrics(
            order_count=int(row.order_count),
            order_sum=_f(row.order_sum),
            buyout_count=int(row.buyout_count),
            buyout_sum=_f(row.buyout_sum),
            buyout_percent=bp,
            returns_count=0,
            cost_price=_f(row.cost_price),
            volume_liters=_f(row.volume_liters),
            first_liter_rate=82.8, extra_liter_rate=25.2,
            warehouse_coef=_f(row.warehouse_coef) or 1.0,
            ktr=1.0, irp=0, commission_pct=3.5,
            seller_coef=0.647, acquiring_pct=2.6,
            logistics_multiplier=1.85,
            ad_spend=_f(row.ad_spend),
            tax_rate=7.0, vat_rate=0,
            stock_quantity=0,
        )
        await db.execute(text("""
            UPDATE product_daily_stats SET
                margin_clean = :mc, margin_clean_pct = :mcp,
                roi = :r, cost_price_total = :cpt
            WHERE id = :id
        """), {
            "mc": round(metrics.get("margin_clean", 0), 2),
            "mcp": round(metrics.get("margin_clean_pct", 0), 2),
            "r": round(metrics.get("roi", 0), 2),
            "cpt": round(metrics.get("cost_price_total", 0), 2),
            "id": row.id,
        })
        updated += 1

    await db.commit()
    return updated


# ======================================================================
# Excel parser
# ======================================================================

async def _read_excel(data: bytes) -> tuple[list[str], list[list]]:
    try:
        import openpyxl
    except ImportError:
        raise HTTPException(500, "openpyxl not installed")

    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.active
    if not ws:
        raise HTTPException(400, "Empty workbook")

    rows_iter = ws.iter_rows(values_only=True)
    header_row = next(rows_iter, None)
    if not header_row:
        raise HTTPException(400, "No header row found")

    headers = [str(h) if h is not None else "" for h in header_row]
    data_rows = [list(r) for r in rows_iter]
    return headers, data_rows


def _excel_to_dicts(headers, data_rows, col_map):
    col_idx = _normalize_headers(headers, col_map)
    required = {"nm_id", "date"} if "retail_price_withdisc_rub" not in col_map.values() else {"nm_id", "rr_dt"}
    detected = set(col_idx.values())

    results = []
    for row in data_rows:
        item = {}
        has_data = False
        for idx, val in enumerate(row):
            mapped = col_idx.get(idx)
            if mapped:
                item[mapped] = val
                if val is not None:
                    has_data = True
        if has_data:
            results.append(item)
    return results, detected


# ======================================================================
# Endpoints
# ======================================================================

@router.post("/upload-stats")
async def upload_stats(file: UploadFile = File(...), db: AsyncSession = Depends(get_db)):
    if not (file.filename or "").endswith(".xlsx"):
        raise HTTPException(400, "Only .xlsx files accepted")
    data = await file.read()
    if len(data) > 50 * 1024 * 1024:
        raise HTTPException(400, "File too large (max 50MB)")

    # Try custom report format (vendor code based, two sheets)
    custom = await _detect_and_import_report(db, data)
    if custom:
        return {"status": "ok", "file": file.filename, "format": "custom_report", **custom}

    headers, rows = await _read_excel(data)
    dicts, cols_found = _excel_to_dicts(headers, rows, PDS_COLUMN_MAP)

    if "nm_id" in cols_found and "date" in cols_found:
        has_order_data = any(d.get("order_count") or d.get("order_sum") or d.get("ad_spend") for d in dicts[:10])
        if has_order_data or any(k in cols_found for k in ("order_count", "order_sum", "ad_spend")):
            result = await _import_pds(db, dicts)
            return {"status": "ok", "file": file.filename, "format": "stats", **result}

    dicts2, cols_found2 = _excel_to_dicts(headers, rows, REAL_COLUMN_MAP)
    if "nm_id" in cols_found2 and "rr_dt" in cols_found2:
        has_real_data = any(d.get("supplier_oper_name") for d in dicts2[:10])
        if has_real_data:
            result = await _import_realization(db, dicts2)
            return {"status": "ok", "file": file.filename, "format": "realization", **result}

    raise HTTPException(400, f"Unknown format. Headers: {headers[:10]}")


@router.post("/upload-realization")
async def upload_realization(file: UploadFile = File(...), db: AsyncSession = Depends(get_db)):
    if not (file.filename or "").endswith(".xlsx"):
        raise HTTPException(400, "Only .xlsx files accepted")
    data = await file.read()
    if len(data) > 50 * 1024 * 1024:
        raise HTTPException(400, "File too large (max 50MB)")

    headers, rows = await _read_excel(data)
    dicts, cols_found = _excel_to_dicts(headers, rows, REAL_COLUMN_MAP)
    if "nm_id" not in cols_found or "rr_dt" not in cols_found:
        raise HTTPException(400, "Required columns (nm_id, date) not found")

    result = await _import_realization(db, dicts)
    return {"status": "ok", "file": file.filename, "format": "realization", **result}
